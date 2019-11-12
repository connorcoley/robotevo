# Copyright (C) 2014-2019, Ariel Vina Rodriguez ( arielvina@yahoo.es )
#  distributed under the GNU General Public License, see <http://www.gnu.org/licenses/>.
#
# author Ariel Vina-Rodriguez (qPCR4vir)
# 2014-2019
__author__ = 'qPCR4vir'

import logging
from pathlib import Path

import EvoScriPy.labware as lab

def_reagent_excess = 4
def_mix_excess = 8


class Reagent:
    """
    A Reagent is a fundamental concept in RobotEvo programming.
    It makes possible to define a protocol in a natural way, matching what a normal
    laboratory's protocol indicates. Defines a named homogeneous liquid solution,
    the wells it occupy, the initial amount needed to run the protocol (auto calculated),
    and how much is needed per sample, if applicable. It is also used to define samples,
    intermediate reactions and products. It makes possible a robust tracking
    of all actions and a logical error detection, while significantly simplifying
    the programming of non trivial protocols.
    """

    current_protocol = None                                               # to (auto)register a list of reagents
    use_minimal_number_of_aliquots = True

    def __init__(self,
                 name           : str,
                 labware        : (lab.Labware, str)        = None,
                 volpersample   : float                     = 0.0,
                 single_use     : float                     = None,
                 wells          : (int, [int], [lab.Well])  = None,
                 replicas       : int                       = None,
                 def_liq_class    : (str,(str,str))           = None,
                 excess         : float                     = None,
                 initial_vol    : float                     = 0.0,
                 maxFull        : float                     = None,
                 num_of_samples : int                       = None,
                 minimize_aliquots : bool                   = None):
        """
        This is a named set of aliquots of an homogeneous solution.
        Put a reagent into labware wells, possible with replicates and set the amount to be used for each sample,
        if applicable.
        This reagent is automatically added to the list of reagents of the worktable were the labware is.
        The specified excess in % will be calculated/expected. A default excess of 4% will be assumed
        if not explicitly indicated.
        A minimal volume will be calculated based on either the number of samples
        and the volume per sample to use or the volume per single use.
        A minimal number of replicas (wells, aliquots) will be calculated based on the minimal volume,
        taking into account the maximum allowed volume per well and the excess specified.

        :param name:            Reagent name. Ex: "Buffer 1", "forward primer", "IC MS2"
        :param labware:         Labware or his label in the worktable; if None will be deduced from `wells`.
        :param volpersample:    how much is needed per sample, if applicable, in uL
        :param single_use;      Not a "per sample" multiple use? Set then here the volume for one single use
        :param wells:           or offset to begging to put replica. If None will try to assign consecutive wells
        :param replicas;        def min_num_of_replica(), number of replicas
        :param def_liq_class;     the name of the Liquid class, as it appears in your own EVOware database.
                                instructions.def_liquidClass if None
        :param excess;          in %
        :param initial_vol;     is set for each replica. If default (=None) is calculated als minimum.
        :param maxFull;         maximo allowed volume in % of the wells capacity
        :param num_of_samples;  if None, the number of samples of the current protocol will be assumed
        :param minimize_aliquots;  use minimal number of aliquots? Defaults to `Reagent.use_minimal_number_of_aliquots`,
                                   This default value can be temporally change by setting that global.
        """
        if labware is None:
            if isinstance(wells, lab.Well):
                labware = wells.labware
            elif isinstance(wells, list) and isinstance(wells[0], lab.Well):
                labware = wells[0].labware
        if isinstance(labware, str):
            labware = Reagent.current_protocol.worktable.get_labware(label=labware)
        assert isinstance(labware, lab.Labware), "No labware defined for Reagent " + name

        # add self to the list of reagents of the worktable were the labware is.
        worktable = None
        assert isinstance(labware.location.worktable, lab.WorkTable)                                   # todo temporal
        if (    isinstance(labware,                     lab.Labware)
            and isinstance(labware.location,            lab.WorkTable.Location)
            and isinstance(labware.location.worktable,  lab.WorkTable)):

            worktable = labware.location.worktable
        else:
            if (Reagent.current_protocol):
                worktable = Reagent.current_protocol.worktable                               # todo temporal

        assert name not in worktable.reagents
        worktable.reagents[name] = self

        ex= def_reagent_excess if excess is None else excess

        self.labware    = labware
        self.maxFull    = 1.0 if maxFull is None else maxFull/100.0
        self.excess     = 1.0 + ex/100.0
        self.def_liq_class = def_liq_class
        self.name       = name
        self.volpersample = volpersample
        self.components = []                                                            # todo reserved for future use
        if minimize_aliquots is not None:
            self.minimize_aliquots = minimize_aliquots
        else:
            self.minimize_aliquots = Reagent.use_minimal_number_of_aliquots
        if single_use:
            assert not volpersample, str(name) + \
                                     ": this is a single use-reagent. Please, don't set any volume per sample."
            if num_of_samples is None:
                num_of_samples = 1
            assert num_of_samples == 1, "this is a single use-reagent, don't set num_of_samples " + str(num_of_samples)

            self.volpersample = single_use
        else:
            if num_of_samples is None:
                num_of_samples = Reagent.current_protocol.num_of_samples or 0

        self.minNumRep = self.min_num_of_replica(NumSamples=num_of_samples)

        # coordinate: minNumRep, wells, replicas, initial_vol. todo fine-tune warnings.
        if replicas is None:
            replicas = self.minNumRep
        assert replicas >= self.minNumRep, ("too few wells (" + str(len(wells)) + ") given for the minimum "
                                                + "number of replica needed (" + str(self.minNumRep) + " )" )

        if isinstance(wells, list):                         # use len(pos) to correct replicas
            assert len(wells) >= self.minNumRep, ("too few wells (" + str(len(wells)) + ") given for the minimum "
                                                + "number of replica needed (" + str(self.minNumRep) + " )" )
            if replicas < len(wells):
                logging.warning("WARNING !! putting more replica of " + name + " to fit the initial volume list provided")
                replicas = len(wells)
            assert not replicas > len(wells), ("too few wells (" + str(len(wells)) + ") given for the desired "
                                                + "number of replicas (" + str(replicas) + " )" )
            if isinstance(initial_vol, list):
                assert replicas == len(initial_vol)
            if replicas > self.minNumRep:                                    # todo revise !! for preMix components
                logging.warning("WARNING !! You may be putting more wells replicas (" + str(len(wells)) + ") of " + name
                      + " that the minimum you need("  + str(self.minNumRep) + " )")

        if not isinstance(initial_vol, list):
            initial_vol = [initial_vol]*replicas

        if replicas < len(initial_vol):
            logging.warning("WARNING !! putting more replica of " + name + " to fit the initial volume list provided")
            replicas = len(initial_vol)
            if replicas > self.minNumRep:                                      # todo revise !! for preMix components
                logging.warning("WARNING !! You may be putting more inital volumens values (" + str(replicas) + ") of " + name
                      + " that the minimum number of replicas you need("  + str(self.minNumRep) + " )")

        self.Replicas   = labware.put(self, wells, self.minNumRep if self.minimize_aliquots else replicas)
        self.pos        = self.Replicas[0].offset                                   # ??

        self.init_vol(NumSamples=num_of_samples, initial_vol=initial_vol)            # put the minimal initial volume
        self.include_in_check = True

    def min_num_of_replica(self, NumSamples: int=None)->int:
        """
        A minimal number of replicas (wells, aliquots) will be calculated based on the minimal volume,
        taking into account the maximum allowed volume per well and the excess specified.
        :param NumSamples:
        :return:
        """
        return int (self.min_vol(NumSamples) / (self.labware.type.max_vol * self.maxFull)) + 1

    @staticmethod
    def set_reagent_list(protocol):
        Reagent.current_protocol = protocol                    # ??

    @staticmethod
    def StopReagentList():
        Reagent.current_protocol = None

    def __str__(self):
        return "{name:s}".format(name=self.name)

    def min_vol(self, NumSamples=None)->float:
        """
        A minimal volume will be calculated based on either the number of samples
        and the volume per sample to use (todo or the volume per single use.)???
        :param NumSamples:
        :return:
        """
        NumSamples = NumSamples or Reagent.current_protocol.num_of_samples or 0
        return self.volpersample * NumSamples * self.excess

    def init_vol(self, NumSamples=None, initial_vol=None):
        if initial_vol is not None:
            assert isinstance(initial_vol, list)
            for w,v in zip(self.Replicas, initial_vol):
                w.vol = v
                assert w.labware.type.max_vol >= w.vol, 'Excess initial volume for '+ str(w)
        self.put_min_vol(NumSamples)

    def put_min_vol(self, NumSamples=None):          # todo create replicas if needed !!!!
        """
        Force you to put an initial volume of reagent that can be used to distribute into samples,
        aspiring equal number of complete doses for each sample from each replica,
        exept the firsts replicas that can be used to aspirate one more dose for the last/rest of samples.
        That is: all replica have equal volumen (number) of doses or the firsts have one more dose
        :param NumSamples:
        :return:
        """
        if NumSamples is None:
            NumSamples = Reagent.current_protocol.num_of_samples
        if not NumSamples:
            return
        V_per_sample = self.volpersample * self.excess
        replicas=len(self.Replicas)
        for i, w in enumerate(self.Replicas):
            v = V_per_sample * (NumSamples + replicas - (i+1))//replicas
            if v > w.vol:  w.vol += (v-w.vol)
            assert w.labware.type.max_vol >= w.vol, 'Add one more replica for '+ w.reagent.name

    def autoselect(self, maxTips=1, offset=None, replicas = None):
        # todo revise !!!!!!!!! we know the wells = Replicas
        return self.labware.autoselect(offset or self.pos, maxTips, len(self.Replicas) if offset is None else 1)

    def select_all(self):
        return self.labware.selectOnly([w.offset for w in self.Replicas])


class Reaction(Reagent):
    def __init__(self, name, track_sample, labware,
                 pos=None, replicas=None, def_liq_class=None, excess=None, initial_vol=None):
        Reagent.__init__(self, name, labware, single_use=0,
                         wells=pos, replicas=replicas, def_liq_class=def_liq_class, excess=excess, initial_vol=initial_vol)
        self.track_sample = track_sample


class preMix(Reagent):

    def __init__(self,
                 name,
                 labware,
                 components,
                 pos        =None,
                 replicas   =None,
                 initial_vol=None,
                 def_liq_class=None,
                 excess     =None,
                 maxFull    =None,
                 num_of_samples = None                  ):

        ex= def_mix_excess if excess is None else excess
        vol=0.0
        for reagent in components:
            vol += reagent.volpersample
            reagent.excess +=  ex/100.0      # todo revise! best to calculate at the moment of making?
            reagent.put_min_vol(num_of_samples)

        if initial_vol is None: initial_vol = 0.0

        Reagent.__init__(self, name,
                         labware,
                         vol,
                         wells= pos,
                         replicas      = replicas,
                         def_liq_class   = def_liq_class,
                         excess        = ex,
                         initial_vol   = initial_vol,
                         maxFull       = maxFull,
                         num_of_samples = num_of_samples)

        self.components = components
        #self.init_vol()

    def init_vol(self, NumSamples=None, initial_vol=None):
        if self.components:
            self.volpersample = 0
            for reagent in self.components:
                self.volpersample += reagent.volpersample
            Reagent.init_vol(self, NumSamples=0, initial_vol=initial_vol)
        pass
        # put volume in replicas only at the moment of making  !!
        # Reagent.init_vol(self, num_samples)
        # self.put_min_vol(num_samples)


    def make(self, NumSamples=None):      # todo deprecate?
        if self.Replicas[0].vol is None:   # ????
            self.put_min_vol(NumSamples)
            assert False
        from EvoScriPy.protocol_steps import makePreMix
        makePreMix(self, NumSamples)

    def compVol(self,index, NumSamples=None):
        NumSamples = NumSamples or Reagent.current_protocol.num_of_samples
        return self.components[index].min_vol(NumSamples)


class Primer:
    ids = {}
    seqs = {}
    names = {}
    key_words = {}
    ids_synt = {}

    next_internal_id = 0

    def __init__(self,
                 name,
                 seq,
                 proposed_stock_conc=100,  # uM
                 id=None,
                 mass=None,  # ug
                 moles=None,  # nmoles
                 molec_w=None,  # g/mol
                 mod_5p=None,
                 mod_3p=None,
                 id_synt=None,
                 kws=None):
        """

        :type moles: float
        """
        self.proposed_stock_conc = proposed_stock_conc
        self.mass = mass
        self.moles = moles
        self.molec_w = molec_w
        self.mod_5p = mod_5p
        self.mod_3p = mod_3p
        self.id_synt = id_synt
        self.name = name
        self.seq = seq
        self.id = id
        self._internal_id = Primer.next_internal_id
        Primer.next_internal_id += 1
        Primer.names.setdefault(name, []).append(self)
        Primer.ids_synt.setdefault(id_synt, []).append(self)

        Primer.ids.setdefault(id, []).append(self)
        Primer.seqs.setdefault(seq, []).append(self)
        if isinstance(kws, list):
            for kw in kws:
                Primer.key_words.setdefault(kw, []).append(self)

    def __str__(self):
        return self.name

    def __repr__(self):
        return (self.name or '-') + '[' + str(self.id or '-') + ']'

    @staticmethod
    def load_excel_list(file_name: Path = None):
        col = {'conc': 0,
               'id': 2,
               'name': 5,
               'moles': 7,
               'mass': 8,
               'seq': 15,
               'mol_w': 11,
               'mod_5p': 19,
               'mod_3p': 20,
               'ido': 22,
               'virus': 25
               }
        logging.debug("opening excel")
        import openpyxl

        if not file_name:
            file_name = Path('C:\Prog\exp\PCR fli.xlsx')

        if not file_name:
            from tkinter import filedialog
            file_name = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsm"), ("All files", "*.*")),
                                                   defaultextension='fas',
                                                   title='Select HEV isolate subtyping deta')
            if not file_name:
                return

        logging.debug(file_name)

        wb = openpyxl.load_workbook(str(file_name))
        logging.debug(wb.sheetnames)

        ws = wb['PCR fli-oligos']
        first = True
        p = None
        for r in ws.iter_rows() :
            if first:
                first = False
            else:
                p=Primer(name=r[col['name']].value,
                         seq=r[col['seq']].value,
                         proposed_stock_conc=r[col['conc']].value,
                         id=r[col['id']].value,
                         mass=r[col['mass']].value,
                         moles=r[col['moles']].value,
                         molec_w=r[col['mol_w']].value,
                         mod_5p=r[col['mod_5p']].value,
                         mod_3p=r[col['mod_3p']].value,
                         id_synt=r[col['ido']].value,
                         kws=[r[col['virus']].value]
                         )
        return p


class PrimerReagent (Reagent):
    excess = def_mix_excess

    def __init__(self,
                 primer: Primer,
                 labware,
                 pos,
                 initial_vol,
                 PCR_conc=0.8,
                 stk_conc=100,
                 excess=None):

        Reagent.__init__(self,
                         primer.name,
                         labware or Lab.stock,
                         initial_vol=initial_vol,
                         excess=Primer.excess)


class PrimerMix:
    ids = {}
    names = {}
    key_words = {}

    next_internal_id = 0

    def __init__(self,
                 name,
                 id = None,
                 conc = 10.0,
                 components = None,
                 ref_vol = None,
                 diluent = None,
                 kws = None):

        self.diluent = diluent
        self.name = name
        self.id = id
        self.conc = conc
        self.components = components
        self.ref_vol = ref_vol

        self._internal_id = PrimerMix.next_internal_id
        PrimerMix.next_internal_id += 1

        PrimerMix.names.setdefault(name, []).append(self)
        PrimerMix.ids.setdefault(id, []).append(self)
        if isinstance(kws, list):
            for kw in kws:
                PrimerMix.key_words.setdefault(kw, []).append(self)

    def __str__(self):
        return self.name

    def __repr__(self):
        return (self.name or '-') + '[' + str(self.id or '-') + ']'

    @staticmethod
    def load_excel_list(file_name: Path = None):
        col = {'conc': 2,
               'id': 0,
               'name': 1,
               'vol': 2,
               'final': 6,
               'virus': 13
               }
        logging.debug("opening excel")
        import openpyxl

        if not file_name:
            file_name = Path('C:\Prog\exp\PCR fli.xlsx')

        if not file_name:
            from tkinter import filedialog
            file_name = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsm"), ("All files", "*.*")),
                                                   defaultextension='fas',
                                                   title='Select HEV isolate subtyping deta')
            if not file_name:
                return

        logging.debug(file_name)

        wb = openpyxl.load_workbook(str(file_name))
        logging.debug(wb.sheetnames)

        ws = wb['Primer mix']
        no_l, header, name_l, vol_l, sep_l, table_h_l, primer_l, diluter_l = 0, 1, 2, 3, 4, 5, 6, 7
        diluter = 'TE 0,1 x'
        line = no_l
        pmix = None
        id = None
        name = None
        conc = None
        ref_vol = None
        kws = None
        components = []

        for r in ws.iter_rows():

            if line == no_l:
                line += 1

            elif line == header:
                id = r[col['id']].value
                kws = [r[col['virus']].value]
                line += 1

            elif line == name_l:
                # assert id == r[col['id']].value
                name = r[col['name']].value
                conc = r[col['conc']].value
                line += 1

            elif line == vol_l:
                ref_vol = r[col['vol']].value
                line += 1

            elif line == sep_l:
                line += 1

            elif line == table_h_l:
                line += 1

            elif line == primer_l:
                comp_name = r[col['name']].value
                if comp_name == diluter:
                    pmix = PrimerMix(name=name,
                                     id=id,
                                     conc=conc,
                                     ref_vol=ref_vol,
                                     kws=kws,
                                     components=components
                                     )
                    line = no_l
                    id = None
                    name = None
                    conc = None
                    ref_vol = None
                    kws = None
                    components = []
                else:
                    components += [(r[col['id']].value,
                                    comp_name,
                                    r[col['conc']].value,
                                    r[col['final']].value)]

        return pmix


class PrimerMixReagent(preMix):
    excess = def_mix_excess

    def __init__(self,
                 name,
                 labware,
                 ID=None,
                 conc=10.0,
                 pos=None,
                 components=None,
                 replicas=1,
                 initial_vol=None,
                 excess=None):

        preMix.__init__(self,
                        name,
                        labware or Lab.stock,
                        pos,
                        components,
                        replicas=replicas,
                        initial_vol=initial_vol,
                        excess=excess or PrimerMix.excess)

        vol=0.0
        for reagent in components:
            vol += reagent.volpersample
            reagent.excess += excess/100.0      # todo revise! best to calculate at the moment of making?
            reagent.put_min_vol()

        if initial_vol is None: initial_vol = 0.0

        Reagent.__init__(self, name, labware, vol, pos=pos, replicas=replicas,
                         defLiqClass=None, excess=ex, initial_vol=initial_vol)
        self.components = components
        #self.init_vol()


class PCRMasterMix(preMix):
    IDs={}
    Names={}
    KWs={}
    Excess = def_mix_excess


    def __init__(self,
                 name,
                 labware,
                 conc       =10.0,
                 pos        =None,
                 components =None,
                 replicas   =1,
                 initial_vol=None,
                 excess     =None):

        preMix.__init__(self, name, labware or Lab.stock, pos, components, replicas=replicas, initial_vol=initial_vol, excess=excess or PrimerMix.Excess)
        vol=0.0
        for reagent in components:
            vol += reagent.volpersample
            reagent.excess +=  ex/100.0      # todo revise! best to calculate at the moment of making?
            reagent.put_min_vol()

        if initial_vol is None: initial_vol = 0.0

        Reagent.__init__(self, name, labware, vol, pos=pos, replicas=replicas,
                         defLiqClass=None, excess=ex, initial_vol=initial_vol)
        self.components = components
        #self.init_vol()


class PCRMix (preMix):
    pass


class PCReaction (Reaction):
    vol = 25
    vol_sample = 5
    pass


class PCRexperiment:
    def __init__(self, ID, name):
        self.PCReactions = []   # list of PCRReaction to create
        self.PCRmixes = []
        self.samples = []
        self.vol = PCReaction.vol
        self.vol_sample = PCReaction.vol_sample

    def addReactions(self, PCRmix: PCRMix, samples, labware: object) -> list:  # of PCRReaction
        pass
    def pippete_mix(self):
        pass
    def pippete_samples(self):
        pass
    def vol (self, vol, vol_sample):
        self.vol = vol
        self.vol_sample = vol_sample



if __name__ == '__main__':
    logging.getLogger(__name__).setLevel(10)
    Primer.load_excel_list()
    PrimerMix.load_excel_list()

